"""
DMP data loader — reads directly from Excel file.
Handles: 173 columns, 179 project rows, multi-line contract clause text in cells.
"""
import pandas as pd
import openpyxl
from pathlib import Path
from utils.helpers import safe_float, parse_date


def load_dmp(filepath: str = None) -> pd.DataFrame:
    """
    Load DMP data from Excel file.

    Args:
        filepath: Path to the DMP Excel file.
                  Defaults to standard location.

    Returns:
        DataFrame with 173 standardized columns, approximately 179 rows.
    """
    if filepath is None:
        base = Path(__file__).parent.parent
        candidates = [
            base / "签约报量（四局）_2026-06-01 13_14_55.xlsx",
            base / "DMP系统市场部字段汇总.xlsx",
        ]
        existing = next((path for path in candidates if path.exists()), None)
        filepath = str(existing) if existing else str(candidates[0])

    # Read directly from Excel to avoid text-extract parsing issues
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Find the main data sheet (签约报量)
    sheet_name = None
    for name in wb.sheetnames:
        if "签约" in name or "报量" in name:
            sheet_name = name
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    # Read all rows
    all_raw = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        all_raw.append(list(row))

    if len(all_raw) < 2:
        return pd.DataFrame()

    # Find transpose marker row (转成列) — truncate before it
    transpose_idx = len(all_raw)
    for i, row in enumerate(all_raw):
        first_cell = str(row[0]) if row[0] is not None else ""
        if "转成列" in first_cell or "转换成列" in first_cell:
            transpose_idx = i
            break

    all_rows = all_raw[:transpose_idx]

    # First row is header
    header = [str(c).strip() if c else "" for c in all_rows[0]]
    data_rows = all_rows[1:]

    # Pad all rows to header length
    ncols = len(header)
    normalized = []
    for row in data_rows:
        row = list(row) if row else []
        if len(row) < ncols:
            row = row + [None] * (ncols - len(row))
        elif len(row) > ncols:
            row = row[:ncols]
        normalized.append(row)

    df = pd.DataFrame(normalized, columns=header)

    # Remove rows without project code
    if "项目编码" in df.columns:
        df = df[df["项目编码"].notna() & (df["项目编码"].astype(str).str.strip() != "")]
    else:
        # Fallback: remove rows where all cells are None
        df = df.dropna(how="all")

    df = df.reset_index(drop=True)

    # Convert numeric columns
    numeric_cols = [
        "签约额（元）", "自行施工金额（元）", "补充协议金额（元）",
        "目标效益率（%）（C值）", "管理效益率（%）（B值）", "一次性经营效益率（%）（A值）",
        "月进度付款比例（%）", "非现金支付比例(%)", "垫资比例(%)",
        "投标担保金额（万元）", "履约担保金额（万元）", "履约担保比例（%）",
        "预付款比例（%）", "质保金支付比例(%)",
        "最高投标限价（元）", "预估垫资金额（万元）",
        "项目投资额（万元）", "项目建安合同额（万元）", "预计自行施工造价（万元）",
        "主体完成付款比例（封顶、形象）(%)", "竣工验收支付比例(%)", "结算支付比例（%）",
        # v2.9: 签约报量(四局)新增字段
        "累计垫资周期", "付款周期",
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = df[col].apply(safe_float)

    # Convert date columns
    date_cols = ["开工时间", "竣工时间", "中标时间", "签约时间", "签约报量时间", "交标时间"]
    for col in date_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    return df


def load_consortium(filepath: str = None) -> pd.DataFrame:
    """Load consortium bidding info from DMP Excel (second sheet)."""
    if filepath is None:
        base = Path(__file__).parent.parent
        filepath = str(base / "DMP系统市场部字段汇总.xlsx")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    # Find the consortium sheet
    for name in wb.sheetnames:
        if "联合体" in name:
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return pd.DataFrame()
            header = [str(c).strip() if c else "" for c in rows[0]]
            data = [list(r) for r in rows[1:]]
            return pd.DataFrame(data, columns=header)

    return pd.DataFrame()
