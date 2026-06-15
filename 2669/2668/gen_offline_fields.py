import openpyxl, time
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

hf = Font(name='Arial', bold=True, size=10, color='FFFFFF')
hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
done_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
pend_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
wrap = Alignment(vertical='center', wrap_text=True)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Sheet 1
ws = wb.active
ws.title = 'offline_fields'

headers = ['Source', 'Owner', 'Field', 'Col/Note', 'Models', 'Purpose', 'Priority', 'Status']
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hf; cell.fill = hfill; cell.border = border; cell.alignment = center

fields = [
    # ====================================================================
    # App1 — 审计附表 (audit working paper)
    # ====================================================================
    ('App1', 'Audit(has)', 'proj_code_fin', 'col 2', 'ALL', 'JOIN key', 'required', 'collected'),
    ('App1', 'Audit(has)', 'cost_profit_rate', 'col 64', '2.2', 'A-value deviation', 'high', 'collected'),
    ('App1', 'Audit(has)', 'adv_receivable', 'col --', '2.3', 'advance gap', 'high', 'collected'),
    ('App1', 'Audit(has)', 'adv_received', 'col --', '2.3', 'advance gap', 'high', 'collected'),
    ('App1', 'Audit(has)', 'cash_balance', 'col --', '2.3', 'negative cash flow', 'high', 'collected'),
    ('App1', 'Audit(has)', 'neg_cash_reason', 'col --', '2.3', 'root cause', 'mid', 'collected'),
    # App1 verify filling
    ('App1', 'Audit(verify)', 'project_status', 'col 5', 'Pre+2.2', 'Pre:unstarted+stopped exempt; 2.2:construction+stopped', 'high', 'v2.8 connected'),
    ('App1', 'Audit(verify)', 'actual_output_val', 'col 65', 'Pre+2.2', 'Pre:180d output check; 2.2:construction verify', 'high', 'v2.8 connected'),
    ('App1', 'Audit(verify)', 'cumulative_receipt', 'col 72', 'Pre+2.2+2.3', 'Pre:360d receipt check; 2.2:verify; 2.3:guarantee', 'high', 'v2.8 connected'),
    ('App1', 'Audit(verify)', 'cert_status_detail', 'col 36', '2.4', 'cert missing reason (DMP has Y/N)', 'mid', 'v2.8 connected'),
    ('App1', 'Audit(verify)', 'delay_days', 'col 34', '2.4', 'schedule delay analysis', 'low', 'v2.8 connected'),
    ('App1', 'Audit(verify)', 'receivables_overdue', 'col 78', '2.1', 'post-sign monitoring', 'mid', 'v2.8 connected'),
    ('App1', 'Audit(verify)', 'unstarted_reason', 'col 35', '2.2', 'stopped root cause supplement', 'low', 'v2.8 connected'),

    # ====================================================================
    # App3 — 中标率统计
    # ====================================================================
    ('App3', 'Audit(has)', 'fail_reason', '--', '2.4', 'bid failure analysis', 'mid', 'collected'),
    ('App3', 'Audit(has)', 'win_rate', '--', '3.1', 'win rate <25% alert', 'high', 'collected'),
    ('App3', 'Audit(has)', 'unit_name', '--', '3.1', 'JOIN key', 'required', 'collected'),

    # ====================================================================
    # App6 — 战略客户统计表
    # ====================================================================
    ('App6', 'Audit(has)', 'cust_name', '--', '1.3,3.2', 'customer key', 'required', 'collected'),
    ('App6', 'Audit(has)', 'cust_tier', '--', '1.3,3.2,3.3', 'strategic customer / rating', 'high', 'collected'),
    ('App6', 'Audit(has)', 'maintainer', '--', '1.3', 'management vacuum check', 'high', 'collected'),

    # ====================================================================
    # DMP — 签约报量(四局) 三证字段 (v2.9: 确认DMP已含，无需OA补充)
    # ====================================================================
    ('DMP', 'DMP(has)', 'has_planning_permit', 'col 26', '2.1,2.4', '是否有规划许可证 — 三证之一', 'required', 'collected'),
    ('DMP', 'DMP(has)', 'has_land_use_permit', 'col 27', '2.1,2.4', '是否有建设用地许可证 — 三证之一', 'required', 'collected'),
    ('DMP', 'DMP(has)', 'has_land_own_permit', 'col 28', '2.1,2.4', '是否有土地使用证 — 三证之一', 'required', 'collected'),

    # ====================================================================
    # DMP — 签约报量(四局) 时间线字段 (v2.9: 替代OA审批流时间)
    # 交标时间 / 中标时间 / 签约报量时间 / 签约时间 四节点全覆盖
    # ====================================================================
    ('DMP', 'DMP(has)', 'submit_time', '签约报量col 33', '1.4', '交标时间 — 投标提交节点，替代OA bid_review_pass_time', 'CRITICAL', 'v2.9 connected'),
    ('DMP', 'DMP(has)', 'bid_win_time', '签约报量col 63', '1.4,2.1,2.2,3.1,3.2,3.3', '中标时间 — 中标结果节点', 'CRITICAL', 'v2.9 connected'),
    ('DMP', 'DMP(has)', 'sign_report_time', '签约报量col 62', '1.4,2.1,2.2', '签约报量时间 — 报量节点', 'CRITICAL', 'v2.9 connected'),
    ('DMP', 'DMP(has)', 'sign_time', '签约报量col 64', '1.4,2.1,2.2,main', '签约时间 — 签约完成节点', 'CRITICAL', 'v2.9 connected'),

    # ====================================================================
    # DMP — 签约报量(四局) 新增关键字段 (v2.9)
    # ====================================================================
    ('DMP', 'DMP(has)', 'pay_condition_fail_flag', '签约报量col 56', '2.1', '是否付款条件不达标 — DMP系统标记，需制度交叉校验', 'CRITICAL', 'v2.9 connected'),
    ('DMP', 'DMP(has)', 'is_bid_restricted', '签约报量col 57', '2.1', '是否限制投标风险 — 系统风控标记', 'high', 'v2.9 connected'),
    ('DMP', 'DMP(has)', 'is_supplement_agreement', '签约报量col 58', 'Pre', '是否为补充协议 — 合同聚合用', 'required', 'v2.9 connected'),
    ('DMP', 'DMP(has)', 'contract_nature', '签约报量col 65', '2.2', '合同性质 — 执行/备案/一致', 'high', 'v2.9 connected'),
    ('DMP', 'DMP(has)', 'cumulative_advance_cycle', '签约报量col 144', '2.2,2.3', '累计垫资周期 — 垫资风险量化', 'high', 'v2.9 connected'),
    ('DMP', 'DMP(has)', 'payment_cycle', '签约报量col 169', '2.4', '付款周期 — 资金回流速度评估', 'mid', 'v2.9 connected'),
    ('DMP', 'DMP(has)', 'remit_account_name', '签约报量col 158', '2.3', '汇款账户名称 — 资金安全校验', 'mid', 'v2.9 connected'),
    ('DMP', 'DMP(has)', 'is_urban_renewal', '签约报量col 17', '1.2', '是否城市更新 — 十五五增长点', 'high', 'v2.9 connected'),
    ('DMP', 'DMP(has)', 'urban_renewal_cat1', '签约报量col 18', '1.2', '城市更新一级分类', 'mid', 'v2.9 connected'),
    ('DMP', 'DMP(has)', 'urban_renewal_cat2', '签约报量col 19', '1.2', '城市更新二级分类', 'mid', 'v2.9 connected'),
    ('DMP', 'DMP(has)', 'consortium_lead_type', '签约报量col 47', '2.3', '联合体牵头人性质 — 联合体风险评估', 'mid', 'v2.9 connected'),
    ('DMP', 'DMP(has)', 'cost_consult_unit', '签约报量col 76', '2.4', '造价咨询单位 — 结算公正性参考', 'low', 'v2.9 connected'),
    ('DMP', 'DMP(has)', 'other_fund_source', '签约报量col 40', '2.1', '其他资金来源说明', 'low', 'v2.9 connected'),
    ('DMP', 'DMP(has)', 'is_key_domain', '签约报量col 43', '1.2', '是否重点领域 — 战略布局标签', 'mid', 'v2.9 connected'),
    ('DMP', 'DMP(has)', 'related_mkt_group', '签约报量col 44', '1.3', '关联营销小组 — 协同营销追踪', 'low', 'v2.9 connected'),
    ('DMP', 'DMP(has)', 'bid_doc_pickup_time', '签约报量col 32', '1.4', '领取招标文件时间 — 招文→交标时序校验', 'high', 'v2.9 connected'),

    # ====================================================================
    # DMP — 待补充字段 (DMP系统暂未提供)
    # ====================================================================
    ('DMP', 'DMP(add)', 'need_financing_help', '--', '2.1', 'financing help = forbidden (rule defined)', 'low', 'pending'),

    # ====================================================================
    # SAP — 财务系统 (仍需从SAP获取)
    # ====================================================================
    ('SAP', 'Finance(new)', 'sap_output_val', '--', '2.2', 'conversion rate (higher precision)', 'mid', 'pending'),
    ('SAP', 'Finance(new)', 'sap_receipt_total', '--', '2.1', 'receipt progress monitor', 'mid', 'pending'),
    ('SAP', 'Finance(new)', 'sap_settled_unpaid', '--', '2.1', 'tiered risk monitor', 'mid', 'pending'),
    ('SAP', 'Finance(new)', 'guarantee_due_date', '--', '2.3', 'overdue = now - due date', 'CRITICAL', 'pending'),
    ('SAP', 'Finance(new)', 'guarantee_actual_date', '--', '2.3', 'compare vs due date', 'CRITICAL', 'pending'),
    ('SAP', 'Finance(new)', 'adv_due_date', '--', '2.3', 'overdue = actual - due', 'CRITICAL', 'pending'),
    ('SAP', 'Finance(new)', 'adv_actual_date', '--', '2.3', 'compare vs due date', 'CRITICAL', 'pending'),

    # ====================================================================
    # CustLedger — 客户台账（v2.9：不可获取，DMP中标报量+制度内嵌名单已覆盖维度三需求）

    # ====================================================================
    # OA — 审批流时间 (v2.9: DMP四节点可覆盖80%场景，OA精确时间仍需IT导出)
    # ====================================================================
    ('OA', 'IT(export)', 'bid_review_pass_time', '--', '1.4', '招文评审通过时间 — DMP交标/中标/报量/签约可部分替代', 'CRITICAL', 'pending'),
    ('OA', 'IT(export)', 'project_approval_time', '--', '1.4', '项目立项审批时间 — DMP中标时间可部分替代', 'mid', 'pending'),
    ('OA', 'IT(export)', 'contract_review_time', '--', '1.4', '合同评审通过时间 — DMP签约报量/签约时间可部分替代', 'mid', 'pending'),
    ('OA', 'IT(export)', 'undertaking_approval_level', '--', '2.1', 'red-line project approval penetration', 'mid', 'pending'),

    # ====================================================================
    # Qichacha — 企查查API
    # ====================================================================
    ('Qichacha', 'Audit(API)', 'is_dishonest', '--', '2.1', 'dishonest = forbidden (rule defined)', 'mid', 'pending'),
    ('Qichacha', 'Audit(API)', 'has_major_lawsuit', '--', '2.1', 'lawsuit = forbidden', 'mid', 'pending'),
    ('Qichacha', 'Audit(API)', 'is_bankrupt', '--', '2.1', 'bankrupt = forbidden', 'mid', 'pending'),
]

for r, (src, owner, fname, col, models, purpose, pri, status) in enumerate(fields, 2):
    vals = [src, owner, fname, col, models, purpose, pri, status]
    is_done = 'collected' in status or 'connected' in status
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        if is_done:
            cell.fill = done_fill
            cell.font = Font(name='Arial', size=9, color='006100')
        else:
            cell.fill = pend_fill
            cell.font = Font(name='Arial', size=9, color='BF8F00')
        cell.border = border; cell.alignment = wrap

for i, w in enumerate([12, 16, 24, 10, 16, 44, 12, 18], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# Sheet 2: Summary
ws2 = wb.create_sheet('Summary')
h2 = ['Source', 'Owner', 'Total', 'Collected/Connected', 'Pending']
for c, h in enumerate(h2, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = hf; cell.fill = hfill; cell.border = border; cell.alignment = center

stats = [
    ('App1 (audit working paper)', 'Audit Team', 13, 6, 7),
    ('App3 (win rate)', 'Audit Team', 3, 3, 0),
    ('App6 (strategic customer)', 'Audit Team', 3, 3, 0),
    ('DMP — 三证 (v2.9)', 'DMP(has)', 3, 3, 0),
    ('DMP — 时间线 (v2.9)', 'DMP(has)', 4, 4, 0),
    ('DMP — 新增关键字段 (v2.9)', 'DMP(has)', 16, 16, 0),
    ('DMP — 待补充', 'DMP(add)', 1, 0, 1),
    ('SAP / Finance System', 'Finance Dept', 7, 0, 7),
    ('OA Approval Flow', 'IT Dept', 4, 0, 4),
    ('Qichacha / Tianyancha', 'Audit Team (API)', 3, 0, 3),
]
for r, (src, owner, total, done, pend) in enumerate(stats, 2):
    for c, v in enumerate([src, owner, total, done, pend], 1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.font = Font(name='Arial', size=10, bold=(c==1))
        cell.border = border; cell.alignment = center if c >= 3 else wrap
tr = 2 + len(stats)
ws2.cell(row=tr, column=1, value='TOTAL').font = Font(name='Arial', bold=True, size=10)
ws2.cell(row=tr, column=3, value=sum(s[2] for s in stats)).font = Font(name='Arial', bold=True, size=10)
ws2.cell(row=tr, column=4, value=sum(s[3] for s in stats)).font = Font(name='Arial', size=10, color='006100')
ws2.cell(row=tr, column=5, value=sum(s[4] for s in stats)).font = Font(name='Arial', size=10, color='BF8F00')
for c in range(1, 6):
    ws2.cell(row=tr, column=c).border = border; ws2.cell(row=tr, column=c).alignment = center
ws2.column_dimensions['A'].width = 30; ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 10; ws2.column_dimensions['D'].width = 20; ws2.column_dimensions['E'].width = 14

ts = time.strftime('%H%M%S')
out = f'c:/Users/sasa/Desktop/offline_fields_{ts}.xlsx'
wb.save(out)
print(f'OK: {out}')
print(f'Fields: {len(fields)} total, {sum(1 for f in fields if "collected" in f[7] or "connected" in f[7])} collected, {sum(1 for f in fields if "pending" in f[7])} pending')
