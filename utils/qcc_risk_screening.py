"""
企查查风险排查表分析引擎

输入: 企查查导出的标准风险排查Excel（14个Sheet）
输出: 补充三个字段 — 是否失信被执行人、是否已违约或破产重整、是否与中建存在重大诉讼
"""
from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── 中建关键词库 ───────────────────────────────────────────
ZHONGJIAN_KEYWORDS = [
    "中国建筑", "中建四局", "中建三局", "中建二局", "中建一局",
    "中建五局", "中建六局", "中建七局", "中建八局", "中建海峡",
    "中建", "中建新疆", "中建国际", "中建方程", "中建南方",
    "中建西部", "中建科工", "中建安装", "中建钢构", "中建交通",
    "中建装饰", "中建市政", "中建地下", "中建港务", "中建路桥",
]

# ── 样式 ────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
BOLD_FONT = Font(bold=True, size=10)
RED_FONT = Font(bold=True, color="9C0006", size=10)
NORMAL_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


# ── 辅助函数 ─────────────────────────────────────────────────
def _safe_amount(val) -> float:
    """安全转换金额为float"""
    if pd.isna(val) or val in (None, "-", "", "/"):
        return 0.0
    try:
        return float(str(val).replace(",", "").replace("，", ""))
    except (ValueError, TypeError):
        return 0.0


def _fmt_money(val: float) -> str:
    """格式化金额显示"""
    if val == 0:
        return "0元"
    v = abs(val)
    if v >= 10000:
        return f"{v / 10000:,.1f}万元"
    return f"{v:,.0f}元"


def _parse_all_roles(parties_text, company_name: str):
    """解析当事人文本，返回 (company_roles, zhongjian_roles)"""
    company_roles = []
    zhongjian_roles = []

    if pd.isna(parties_text) or not company_name:
        return company_roles, zhongjian_roles

    for line in str(parties_text).strip().split("\n"):
        line = line.strip()
        if not line:
            continue
        parts = line.split(" - ", 1)
        if len(parts) < 2:
            continue
        role, name = parts[0].strip(), parts[1].strip()

        is_zj = any(kw in name for kw in ZHONGJIAN_KEYWORDS)
        is_corp = company_name in name

        if is_zj:
            zhongjian_roles.append(role)
        if is_corp:
            company_roles.append(role)

    return company_roles, zhongjian_roles


# ── 主分析函数 ───────────────────────────────────────────────
def analyze_qcc_risk(file_path: str, output_dir: str | None = None) -> dict:
    """
    分析企查查风险排查表，补充三个缺失字段。

    Parameters
    ----------
    file_path : str
        企查查导出Excel文件路径
    output_dir : str, optional
        补充后Excel输出目录，默认与输入文件同目录

    Returns
    -------
    dict:
        {
            "companies": [             # 全部企业分析结果
                {
                    "name": str,
                    "is_shixin": bool,
                    "is_weiyue": bool,
                    "is_zhongjian": str,  # "否" | "是（与中建同为被告）" | "是（被中建起诉）" | "是（公司非被告）"
                    "details": [str, ...]
                },
                ...
            ],
            "summary": {
                "total": int,
                "shixin_count": int,
                "weiyue_count": int,
                "zhongjian_count": int,
                "zhongjian_not_defendant": int,
                "zhongjian_as_defendant": int,
            },
            "zhongjian_cases": [...],  # 中建相关案件详情
            "output_path": str,        # 补充后Excel路径
        }
    """
    if output_dir is None:
        output_dir = os.path.dirname(file_path)

    # ── 1. 读取数据（兼容不同Sheet数量的企查查导出） ──
    xls = pd.ExcelFile(file_path)
    available_sheets = set(xls.sheet_names)

    # 必选Sheet
    if "统计" not in available_sheets:
        raise ValueError("缺少必需的Sheet: 统计")
    if "裁判文书" not in available_sheets:
        raise ValueError("缺少必需的Sheet: 裁判文书")

    df_main = pd.read_excel(file_path, sheet_name="统计", header=2)
    df_judge = pd.read_excel(file_path, sheet_name="裁判文书", header=1)

    # 可选Sheet — 不存在则用空DataFrame
    def _safe_read(sheet_name, header_row=1):
        if sheet_name in available_sheets:
            return pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
        return pd.DataFrame()

    df_shixin = _safe_read("失信被执行人")
    df_zhixing = _safe_read("被执行人")
    df_zhongben = _safe_read("终本案件")
    df_xianzhi = _safe_read("限制高消费")
    df_dongjie = _safe_read("股权冻结")

    # 主表中列索引可能不同，动态查找
    main_cols = {str(df_main.columns[i]).strip(): i for i in range(len(df_main.columns))}

    def _main_col(name):
        """获取主表中指定列名的索引，不存在返回 -1"""
        return main_cols.get(name, -1)

    # ── 2. 失信被执行人判定 ──
    shixin_set = set()
    # 方法A: 主表"失信被执行人"列 > 0
    sx_col = _main_col("失信被执行人")
    if sx_col >= 0:
        for _, row in df_main.iterrows():
            val = row.iloc[sx_col]
            if pd.notna(val) and str(val) not in ("0", "0.00", "0.0", "nan", ""):
                try:
                    if float(str(val).replace(",", "")) > 0:
                        shixin_set.add(str(row.iloc[1]).strip())
                except ValueError:
                    pass
    # 方法B: 失信被执行人sheet
    if not df_shixin.empty and "企业名称" in df_shixin.columns:
        for n in df_shixin["企业名称"].dropna():
            shixin_set.add(str(n).strip())

    # ── 3. 违约/破产重整判定 ──
    pochan_set = set()
    pc_col = _main_col("破产重整")
    if pc_col >= 0:
        for _, row in df_main.iterrows():
            val = row.iloc[pc_col]
            if pd.notna(val) and str(val) not in ("0", "0.00", "0.0", "nan", ""):
                try:
                    if float(str(val).replace(",", "")) > 0:
                        pochan_set.add(str(row.iloc[1]).strip())
                except ValueError:
                    pass

    def _col_set(df, col="企业名称"):
        if df.empty or col not in df.columns:
            return set()
        return {str(n).strip() for n in df[col].dropna()}

    zhixing_set = _col_set(df_zhixing)
    zhongben_set = _col_set(df_zhongben)
    xianzhi_set = _col_set(df_xianzhi)
    dongjie_set = _col_set(df_dongjie)

    weiyue_set = pochan_set | zhixing_set | zhongben_set | xianzhi_set | dongjie_set

    # ── 4. 中建重大诉讼判定 ──
    company_zhongjian_cases = {}  # company_name -> [case_dict, ...]
    all_zhongjian_cases = []

    for _, row in df_judge.iterrows():
        parties = row["当事人"] if pd.notna(row["当事人"]) else ""
        company_name = str(row["企业名称"]).strip() if pd.notna(row["企业名称"]) else ""

        matched_kw = None
        for kw in ZHONGJIAN_KEYWORDS:
            if kw in str(parties):
                matched_kw = kw
                break

        if matched_kw:
            company_roles, zhongjian_roles = _parse_all_roles(parties, company_name)

            case_info = {
                "企业名称": company_name,
                "文书标题": str(row["文书标题"]) if pd.notna(row["文书标题"]) else "",
                "案号": str(row["案号"]) if pd.notna(row["案号"]) else "",
                "案由": str(row["案由"]) if pd.notna(row["案由"]) else "",
                "案件金额": str(row["案件金额（元）"]) if pd.notna(row["案件金额（元）"]) else "-",
                "裁判日期": str(row["裁判日期"]) if pd.notna(row["裁判日期"]) else "",
                "匹配关键词": matched_kw,
                "企业角色": ", ".join(company_roles) if company_roles else "未识别",
                "中建方角色": ", ".join(zhongjian_roles) if zhongjian_roles else "未识别",
                "是否同方": bool(set(company_roles) & set(zhongjian_roles)) if company_roles and zhongjian_roles else False,
            }
            all_zhongjian_cases.append(case_info)
            company_zhongjian_cases.setdefault(company_name, []).append(case_info)

    # 分类: 公司非被告 vs 公司为被告
    companies_not_defendant = set()
    companies_as_defendant = set()
    for company, cases in company_zhongjian_cases.items():
        for case in cases:
            is_def = any(r in case["企业角色"] for r in
                         ["被告", "被执行人", "被申请人", "被上诉人",
                          "被申请执行人", "被异议人", "被再审人"])
            if is_def:
                companies_as_defendant.add(company)
            else:
                companies_not_defendant.add(company)

    all_zhongjian = set(company_zhongjian_cases.keys())

    # ── 5. 构建企业分析结果 ──
    companies = []
    for _, row in df_main.iterrows():
        cn = str(row.iloc[1]).strip()
        if not cn or cn == "nan":
            continue

        is_sx = cn in shixin_set
        is_wy = cn in weiyue_set

        # 中建诉讼状态
        if cn in companies_not_defendant:
            zj_status = "是（公司非被告）"
        elif cn in companies_as_defendant:
            same_side = any(c["是否同方"] for c in company_zhongjian_cases.get(cn, []))
            if same_side:
                zj_status = "是（与中建同为被告）"
            else:
                zj_status = "是（被中建起诉）"
        else:
            zj_status = "否"

        # 详情
        details = []
        if is_sx:
            if not df_shixin.empty and "企业名称" in df_shixin.columns and "涉案金额(元)" in df_shixin.columns:
                sx_cases = df_shixin[df_shixin["企业名称"].str.strip() == cn]
                total_amt = sum(_safe_amount(a) for a in sx_cases["涉案金额(元)"])
                details.append(f"【失信被执行人】{len(sx_cases)}条记录，涉案{_fmt_money(total_amt)}")
            else:
                details.append("【失信被执行人】是")

        if cn in pochan_set:
            details.append("【破产重整】涉及破产重整案件")
        if cn in zhixing_set:
            if not df_zhixing.empty:
                zx = df_zhixing[df_zhixing["企业名称"].str.strip() == cn]
                details.append(f"【被执行人】{len(zx)}条被执行记录")
            else:
                details.append("【被执行人】是（主表标记）")
        if cn in zhongben_set:
            if not df_zhongben.empty:
                zb = df_zhongben[df_zhongben["企业名称"].str.strip() == cn]
                details.append(f"【终本案件】{len(zb)}条，执行不能风险")
            else:
                details.append("【终本案件】是")
        if cn in xianzhi_set:
            if not df_xianzhi.empty:
                xz = df_xianzhi[df_xianzhi["企业名称"].str.strip() == cn]
                details.append(f"【限制高消费】{len(xz)}条记录")
            else:
                details.append("【限制高消费】是")
        if cn in dongjie_set:
            if not df_dongjie.empty:
                dj = df_dongjie[df_dongjie["企业名称"].str.strip() == cn]
                details.append(f"【股权冻结】{len(dj)}条冻结记录")
            else:
                details.append("【股权冻结】是")
        if cn in company_zhongjian_cases:
            for case in company_zhongjian_cases[cn]:
                amt = f"，金额{case['案件金额']}" if case["案件金额"] and case["案件金额"] != "-" else ""
                details.append(
                    f"【中建诉讼】{case['案号']}（{case['案由']}）\n"
                    f"  企业角色={case['企业角色']}，中建角色={case['中建方角色']}{amt}"
                )

        companies.append({
            "name": cn,
            "is_shixin": is_sx,
            "is_weiyue": is_wy,
            "is_zhongjian": zj_status,
            "details": details,
        })

    # ── 6. 生成补充后Excel ──
    output_path = os.path.join(
        output_dir,
        f"风险排查_补充分析_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    _generate_output_excel(
        file_path, output_path,
        shixin_set, weiyue_set, pochan_set,
        zhixing_set, zhongben_set, xianzhi_set, dongjie_set,
        companies_not_defendant, companies_as_defendant,
        company_zhongjian_cases, all_zhongjian_cases,
        df_shixin, df_zhixing, df_zhongben, df_xianzhi, df_dongjie,
    )

    # ── 7. 汇总 ──
    total = len(companies)
    summary = {
        "total": total,
        "shixin_count": len(shixin_set),
        "weiyue_count": len(weiyue_set),
        "zhongjian_count": len(all_zhongjian),
        "zhongjian_not_defendant": len(companies_not_defendant),
        "zhongjian_as_defendant": len(companies_as_defendant),
    }

    return {
        "companies": companies,
        "summary": summary,
        "zhongjian_cases": all_zhongjian_cases,
        "output_path": output_path,
    }


def _generate_output_excel(
    src_path, output_path,
    shixin_set, weiyue_set, pochan_set,
    zhixing_set, zhongben_set, xianzhi_set, dongjie_set,
    companies_not_defendant, companies_as_defendant,
    company_zhongjian_cases, all_zhongjian_cases,
    df_shixin, df_zhixing, df_zhongben, df_xianzhi, df_dongjie,
):
    """生成补充后的Excel文件"""
    wb = load_workbook(src_path)
    ws = wb["统计"]
    header_row = 3

    nc1 = ws.max_column + 1
    nc2 = ws.max_column + 2
    nc3 = ws.max_column + 3
    nc4 = ws.max_column + 4

    headers = ["是否失信被执行人", "是否已违约或破产重整", "是否与中建存在重大诉讼", "风险详情说明"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=header_row, column=nc1 + i, value=h)
        cell.fill = HEADER_FILL
        cell.font = BOLD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # 写入数据
    for row in range(header_row + 1, ws.max_row + 1):
        cn_val = ws.cell(row=row, column=2).value
        if not cn_val:
            continue
        cn = str(cn_val).strip()

        is_sx = "是" if cn in shixin_set else "否"
        c1 = ws.cell(row=row, column=nc1, value=is_sx)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.border = THIN_BORDER
        c1.font = NORMAL_FONT
        if is_sx == "是":
            c1.fill = RED_FILL
            c1.font = RED_FONT

        is_wy = "是" if cn in weiyue_set else "否"
        c2 = ws.cell(row=row, column=nc2, value=is_wy)
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = THIN_BORDER
        c2.font = NORMAL_FONT
        if is_wy == "是":
            c2.fill = RED_FILL
            c2.font = RED_FONT

        if cn in companies_not_defendant:
            zj = "是（公司非被告）"
        elif cn in companies_as_defendant:
            same = any(c["是否同方"] for c in company_zhongjian_cases.get(cn, []))
            zj = "是（与中建同为被告）" if same else "是（被中建起诉）"
        else:
            zj = "否"

        c3 = ws.cell(row=row, column=nc3, value=zj)
        c3.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c3.border = THIN_BORDER
        c3.font = NORMAL_FONT
        if "是" in zj:
            c3.fill = RED_FILL
            c3.font = RED_FONT

        # 详情
        details = []
        if cn in shixin_set:
            if not df_shixin.empty and "企业名称" in df_shixin.columns and "涉案金额(元)" in df_shixin.columns:
                sx = df_shixin[df_shixin["企业名称"].str.strip() == cn]
                amt = sum(_safe_amount(a) for a in sx["涉案金额(元)"])
                details.append(f"【失信被执行人】{len(sx)}条，涉案{_fmt_money(amt)}")
            else:
                details.append("【失信被执行人】是")
        if cn in pochan_set:
            details.append("【破产重整】涉及破产重整案件")
        if cn in zhixing_set:
            if not df_zhixing.empty:
                details.append(f"【被执行人】{len(df_zhixing[df_zhixing['企业名称'].str.strip() == cn])}条")
            else:
                details.append("【被执行人】是（主表标记）")
        if cn in zhongben_set:
            if not df_zhongben.empty:
                details.append(f"【终本案件】{len(df_zhongben[df_zhongben['企业名称'].str.strip() == cn])}条")
            else:
                details.append("【终本案件】是")
        if cn in xianzhi_set:
            if not df_xianzhi.empty:
                details.append(f"【限制高消费】{len(df_xianzhi[df_xianzhi['企业名称'].str.strip() == cn])}条")
            else:
                details.append("【限制高消费】是")
        if cn in dongjie_set:
            if not df_dongjie.empty:
                details.append(f"【股权冻结】{len(df_dongjie[df_dongjie['企业名称'].str.strip() == cn])}条")
            else:
                details.append("【股权冻结】是")
        if cn in company_zhongjian_cases:
            for case in company_zhongjian_cases[cn]:
                amt = f"，金额{case['案件金额']}" if case["案件金额"] and case["案件金额"] != "-" else ""
                details.append(
                    f"【中建诉讼】{case['案号']}（{case['案由']}）\n"
                    f"  企业={case['企业角色']}，中建={case['中建方角色']}{amt}"
                )

        c4 = ws.cell(row=row, column=nc4, value="\n".join(details) if details else "无异常")
        c4.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c4.border = THIN_BORDER
        c4.font = NORMAL_FONT

    # 调整列宽
    for ci in [nc1, nc2, nc3, nc4]:
        letter = ws.cell(row=header_row, column=ci).column_letter
        if ci == nc4:
            ws.column_dimensions[letter].width = 65
        elif ci == nc3:
            ws.column_dimensions[letter].width = 28
        else:
            ws.column_dimensions[letter].width = 20

    # 在裁判文书sheet标记中建相关
    if "裁判文书" in wb.sheetnames:
        ws_j = wb["裁判文书"]
        j_header = 2
        mc_col = ws_j.max_column + 1
        mc = ws_j.cell(row=j_header, column=mc_col, value="涉及中建")
        mc.fill = HEADER_FILL
        mc.font = BOLD_FONT
        mc.border = THIN_BORDER

        zj_nums = {c["案号"] for c in all_zhongjian_cases}
        for row in range(j_header + 1, ws_j.max_row + 1):
            case_num = ws_j.cell(row=row, column=5).value
            cell = ws_j.cell(row=row, column=mc_col)
            cell.border = THIN_BORDER
            if case_num and str(case_num).strip() in zj_nums:
                cell.value = "是（中建相关）"
                cell.fill = YELLOW_FILL
                cell.font = Font(bold=True, color="9C0006", size=9)
            else:
                cell.value = "否"
                cell.font = Font(size=9)

    wb.save(output_path)
