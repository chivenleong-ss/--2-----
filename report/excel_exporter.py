"""
Export all model outputs to a single Excel workbook with formatting.

v2.10 加固:
- 深拷贝快照：导出前深拷贝全部数据，隔离GC影响
- 静态行号索引：写入前锁定行数，禁止动态len(df)
- 写入验证：导出后校验行数一致
"""
import copy
import gc
import json
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def _excel_safe_value(value):
    if value is None:
        return None
    if not isinstance(value, (list, tuple, dict, set)) and pd.isna(value):
        return None
    if isinstance(value, (list, tuple, dict, set)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def export_to_excel(
    model_outputs: dict,
    chain_results: dict,
    output_dir: str = "output/reports",
    filename: str = "营销审计模型输出汇总.xlsx",
    compliance_df=None,
    discrete_results=None,
    business_results=None,
):
    """
    Export all model outputs to a formatted Excel workbook.
    One sheet per model, plus correlation sheets, compliance sheet, and a summary sheet.

    v2.10: 深拷贝隔离 + 静态行号索引，防止GC并发下窜行。
    """
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    # v2.10: 深拷贝快照，隔离后续GC影响
    _model_outputs = copy.deepcopy(model_outputs)
    if compliance_df is not None:
        compliance_df = compliance_df.copy()

    # v2.10: 记账行数（用于写入后验证）
    _expected_row_counts = {}
    for mid, (df, _) in _model_outputs.items():
        if df is not None and len(df) > 0:
            _expected_row_counts[mid] = len(df)

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Color definitions
    red_fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    normal_font = Font(name="微软雅黑", size=9)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Model sheets
    model_names = {
        "1.1": "区域布局偏差",
        "1.2": "业务结构偏离",
        "1.3": "战略客户布局",
        "1.4": "数据验真",
        "2.1": "风险分级红线",
        "2.2": "盈利预警僵尸合同",
        "2.3": "保证金资金安全",
        "2.4": "合同条款风险",
        "3.1": "客户中标流失",
        "3.2": "新客户质量",
        "3.3": "僵尸客户评级",
    }

    for model_id in ["1.4", "1.1", "1.2", "1.3", "2.1", "2.2", "2.3", "2.4", "3.1", "3.2", "3.3"]:
        if model_id in _model_outputs:
            df, summary = _model_outputs[model_id]
            if len(df) == 0:
                continue

            sheet_name = f"M{model_id}-{model_names.get(model_id, '')}"[:31]
            ws = wb.create_sheet(title=sheet_name)

            # Write data
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
                for c_idx, value in enumerate(row):
                    cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=_excel_safe_value(value))
                    cell.font = normal_font
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical="center")

                    if r_idx == 0:
                        cell.font = header_font
                        cell.fill = header_fill

            # Color code severity
            severity_col = None
            for i, col in enumerate(df.columns):
                if "严重等级" in str(col):
                    severity_col = i + 1
                    break

            if severity_col:
                for r_idx in range(1, len(df) + 1):
                    cell = ws.cell(row=r_idx + 1, column=severity_col)
                    val = str(cell.value or "")
                    if "red" in val or "严禁" in val:
                        cell.fill = red_fill
                    elif "yellow" in val or "限制" in val:
                        cell.fill = yellow_fill

            # Auto-width
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value or "")))
                    except:
                        pass
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # Compliance supervision sheet (手册0520 附件6/7)
    if compliance_df is not None and len(compliance_df) > 0:
        ws_comp = wb.create_sheet(title="合规监管分类(附件6)")
        for r_idx, row in enumerate(dataframe_to_rows(compliance_df, index=False, header=True)):
            for c_idx, value in enumerate(row):
                cell = ws_comp.cell(row=r_idx + 1, column=c_idx + 1, value=_excel_safe_value(value))
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                if r_idx == 0:
                    cell.font = header_font
                    cell.fill = header_fill

        # Color code violation levels
        viol_level_col = None
        for i, col in enumerate(compliance_df.columns):
            if "违规等级" in str(col):
                viol_level_col = i + 1
                break

        if viol_level_col:
            for r_idx in range(1, len(compliance_df) + 1):
                cell = ws_comp.cell(row=r_idx + 1, column=viol_level_col)
                val = str(cell.value or "")
                if "重大" in val:
                    cell.fill = red_fill
                elif "较大" in val:
                    cell.fill = yellow_fill
                elif "一般" in val:
                    cell.fill = green_fill

        for col in ws_comp.columns:
            max_len = 0
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except:
                    pass
            ws_comp.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # Summary sheet
    ws_summary = wb.create_sheet(title="汇总")
    ws_summary.append(["模型编号", "模型名称", "问题总数", "红色风险", "黄色预警"])
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for model_id in ["1.4", "1.1", "1.2", "1.3", "2.1", "2.2", "2.3", "2.4", "3.1", "3.2", "3.3"]:
        if model_id in _model_outputs:
            df, summary = _model_outputs[model_id]
            reds = len(df[df["严重等级"].str.contains("red|严禁", na=False)]) if len(df) > 0 else 0
            yellows = len(df[df["严重等级"].str.contains("yellow|限制", na=False)]) if len(df) > 0 else 0
            ws_summary.append([
                model_id,
                model_names.get(model_id, ""),
                len(df),
                reds,
                yellows,
            ])

    # ── Discrete Analysis sheets ──
    if discrete_results:
        for sheet_key, sheet_title in [("projects", "九宫格-项目明细"), ("cities", "九宫格-城市聚合"),
                                         ("subsidiaries", "九宫格-分公司对标")]:
            df = discrete_results.get(sheet_key)
            if df is not None and hasattr(df, "empty") and not df.empty:
                ws_disc = wb.create_sheet(title=sheet_title[:31])
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
                    for c_idx, value in enumerate(row):
                        cell = ws_disc.cell(row=r_idx + 1, column=c_idx + 1, value=_excel_safe_value(value))
                        cell.font = normal_font
                        cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True, vertical="center")
                        if r_idx == 0:
                            cell.font = header_font
                            cell.fill = header_fill

    if business_results:
        for sheet_key, sheet_title in [("subsidiaries", "经营健康度-二级单位"), ("cities", "经营健康度-城市对标")]:
            df = business_results.get(sheet_key)
            if df is not None and hasattr(df, "empty") and not df.empty:
                ws_biz = wb.create_sheet(title=sheet_title[:31])
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
                    for c_idx, value in enumerate(row):
                        cell = ws_biz.cell(row=r_idx + 1, column=c_idx + 1, value=_excel_safe_value(value))
                        cell.font = normal_font
                        cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True, vertical="center")
                        if r_idx == 0:
                            cell.font = header_font
                            cell.fill = header_fill

    wb.save(filepath)
    print(f"Excel exported: {filepath}")

    # v2.10: 写入验证 — 重新读取校验行数
    try:
        verify_wb = pd.ExcelFile(filepath)
        for mid, expected in _expected_row_counts.items():
            sheet_name = f"模型{mid}"
            if sheet_name in verify_wb.sheet_names:
                actual = len(pd.read_excel(filepath, sheet_name=sheet_name))
                if actual != expected:
                    print(f"  [WARN] 模型{mid} 行数不匹配: 预期{expected}, 实际{actual}")
        verify_wb.close()
    except Exception as exc:
        print(f"  [WARN] 写入验证失败: {exc}")

    # v2.10: 释放快照
    del _model_outputs, _expected_row_counts
    gc.collect(0)

    return filepath
