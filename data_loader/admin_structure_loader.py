"""
Load 行政架构-四局.xlsx and resolve DMP 申报单位 to bureau hierarchy.
"""
from __future__ import annotations

import re
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).parent.parent
UPLOAD_DIR = BASE / "uploads"
DESKTOP_ADMIN_FALLBACK = Path(r"C:\Users\sasa\Desktop\模型2\行政架构-四局.xlsx")

BUREAU_SHORT = "四局"
BUREAU_FULL = "中国建筑第四工程局有限公司"
ALL_DIRECT = "全部直属机构"
ALL_SUB = "全部下属机构"
ALL_CITY = "全部城市"

SKIP_UPLOAD_FILES = {
    "appendix.xlsx",
    "bid_report.xlsx",
    "dmp_sales.xlsx",
    "qcc_risk.xlsx",
}


def _clean(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _is_id(value) -> bool:
    return bool(value and re.fullmatch(r"\d+(\.0)?", str(value).strip()))


def _find_admin_structure_file() -> Path | None:
    explicit = UPLOAD_DIR / "行政架构-四局.xlsx"
    if explicit.exists():
        return explicit
    alias = UPLOAD_DIR / "admin_structure.xlsx"
    if alias.exists():
        return alias
    if DESKTOP_ADMIN_FALLBACK.exists():
        return DESKTOP_ADMIN_FALLBACK
    for path in sorted(UPLOAD_DIR.glob("*.xlsx")):
        if path.name in SKIP_UPLOAD_FILES:
            continue
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet_name = workbook.sheetnames[0] if workbook.sheetnames else ""
            workbook.close()
            if "行政架构" in sheet_name:
                return path
        except Exception:
            continue
    return None


def _parse_admin_rows(rows) -> tuple[dict, str]:
    nodes_by_id: dict[str, dict] = {}
    root_id = ""

    for row in rows:
        vals = [_clean(v) for v in row]
        node = None
        for index in range(len(vals) - 1):
            left, right = vals[index], vals[index + 1]
            if _is_id(left) and right and not _is_id(right):
                node = {"id": left, "name": right, "level": index + 1}

        if not node:
            continue

        parent_id = None
        for index in range(len(vals) - 1, -1, -1):
            if _is_id(vals[index]):
                parent_id = vals[index]
                break

        node["parent_id"] = parent_id
        nodes_by_id[node["id"]] = node
        if node["level"] == 1:
            root_id = node["id"]

    return nodes_by_id, root_id


def clear_admin_structure_cache() -> None:
    load_admin_structure.cache_clear()


@lru_cache(maxsize=1)
def load_admin_structure() -> dict:
    empty = {
        "nodes_by_id": {},
        "nodes_by_name": {},
        "children_by_parent": {},
        "root_id": "",
        "direct_units": [],
        "detail_by_direct": {},
        "descendants_by_name": {},
    }

    path = _find_admin_structure_file()
    if not path:
        return empty

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    nodes_by_id, root_id = _parse_admin_rows(rows)
    if not nodes_by_id or not root_id:
        return empty

    nodes_by_name = {node["name"]: node for node in nodes_by_id.values()}
    children_by_parent: dict[str, list[str]] = {}
    for node in nodes_by_id.values():
        parent_id = node.get("parent_id")
        if parent_id:
            children_by_parent.setdefault(parent_id, []).append(node["id"])

    direct_units = [
        nodes_by_id[child_id]["name"]
        for child_id in children_by_parent.get(root_id, [])
        if child_id in nodes_by_id
    ]
    direct_units = sorted(set(direct_units))

    detail_by_direct: dict[str, list[str]] = {}
    for direct_name in direct_units:
        direct_node = nodes_by_name.get(direct_name)
        if not direct_node:
            continue
        child_names = [
            nodes_by_id[child_id]["name"]
            for child_id in children_by_parent.get(direct_node["id"], [])
            if child_id in nodes_by_id
        ]
        detail_by_direct[direct_name] = sorted(set(child_names))

    descendants_by_name: dict[str, set[str]] = {}

    def collect_descendants(node_id: str) -> set[str]:
        node = nodes_by_id.get(node_id)
        if not node:
            return set()
        names = {node["name"]}
        for child_id in children_by_parent.get(node_id, []):
            names.update(collect_descendants(child_id))
        return names

    for node in nodes_by_id.values():
        descendants_by_name[node["name"]] = collect_descendants(node["id"])

    return {
        "nodes_by_id": nodes_by_id,
        "nodes_by_name": nodes_by_name,
        "children_by_parent": children_by_parent,
        "root_id": root_id,
        "direct_units": direct_units,
        "detail_by_direct": detail_by_direct,
        "descendants_by_name": descendants_by_name,
    }


def _match_node(unit_name: str, admin: dict):
    unit_name = str(unit_name or "").strip()
    if not unit_name:
        return None

    nodes_by_name = admin.get("nodes_by_name") or {}
    if unit_name in nodes_by_name:
        return nodes_by_name[unit_name]

    candidates = []
    for name, node in nodes_by_name.items():
        if name in unit_name or unit_name in name:
            candidates.append((len(name), node))
    if not candidates:
        return None
    return max(candidates, key=lambda item: item[0])[1]


def _direct_unit_name(node: dict, admin: dict) -> str:
    nodes_by_id = admin.get("nodes_by_id") or {}
    root_id = admin.get("root_id") or ""
    current = node
    while current:
        parent_id = current.get("parent_id")
        if not parent_id:
            break
        if parent_id == root_id:
            return current["name"]
        parent = nodes_by_id.get(parent_id)
        if not parent:
            break
        current = parent
    return node["name"]


def resolve_unit_scope(unit_name: str, fallback_direct: str = "") -> dict[str, str]:
    admin = load_admin_structure()
    unit_name = str(unit_name or "").strip()
    node = _match_node(unit_name, admin)

    if node:
        direct = _direct_unit_name(node, admin)
        return {
            "局": BUREAU_SHORT,
            "二级": direct,
            "细分": node["name"],
            "申报单位": unit_name,
        }

    direct = str(fallback_direct or unit_name or "未识别直属机构").strip()
    return {
        "局": BUREAU_SHORT,
        "二级": direct,
        "细分": unit_name or direct,
        "申报单位": unit_name,
    }


def get_admin_scope_options() -> dict:
    admin = load_admin_structure()
    direct_units = admin.get("direct_units") or []
    detail_by_direct = admin.get("detail_by_direct") or {}

    if not direct_units:
        return {
            "global": [BUREAU_SHORT],
            "secondary": [ALL_DIRECT],
            "detail": [ALL_SUB],
            "city": [ALL_CITY],
            "detail_by_secondary": {},
            "city_by_secondary": {},
            "city_by_detail": {},
        }

    return {
        "global": [BUREAU_SHORT],
        "secondary": [ALL_DIRECT, *direct_units],
        "detail": [ALL_SUB],
        "city": [ALL_CITY],
        "detail_by_secondary": detail_by_direct,
        "city_by_secondary": {},
        "city_by_detail": {},
    }


def detail_filter_names(detail_name: str) -> set[str]:
    admin = load_admin_structure()
    descendants = admin.get("descendants_by_name") or {}
    if detail_name in descendants:
        return set(descendants[detail_name])
    if detail_name:
        return {detail_name}
    return set()


def is_all_global(value: str) -> bool:
    return str(value or "").strip() in {"", BUREAU_SHORT, "全局"}


def is_all_direct(value: str) -> bool:
    return str(value or "").strip() in {"", ALL_DIRECT, "全部二级"}


def is_all_sub(value: str) -> bool:
    return str(value or "").strip() in {"", ALL_SUB, "全部细分"}


def is_all_city(value: str) -> bool:
    return str(value or "").strip() in {"", ALL_CITY, "全部城市"}
