"""
Loaders for audit appendix tables.

Supports two workbook layouts:
1. Legacy multi-sheet appendix workbook.
2. New merged workbook where Appendix 1 already absorbs SAP-finance fields.
"""
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

from utils.helpers import safe_float

BASE = Path(__file__).parent.parent
UPLOADS_DIR = BASE / "uploads"
APPENDIX_UPLOAD_CANDIDATES = [
    UPLOADS_DIR / "appendix.xlsx",
    UPLOADS_DIR / "appendix.xls",
    BASE / "【附件5】附表：营销质量管理专项审计附表（定稿）(1).xlsx",
    BASE / "【725】附件：营销质量管理专项审计附表（定稿）(1).xlsx",
]
APPENDIX_EXCEL = str(APPENDIX_UPLOAD_CANDIDATES[-1])


def _resolve_appendix_excel_path(filepath: str | None = None) -> str:
    if filepath:
        return filepath
    for candidate in APPENDIX_UPLOAD_CANDIDATES:
        if candidate.exists():
            return str(candidate)
    return APPENDIX_EXCEL

MERGED_APPENDIX_REQUIRED_COLUMNS = [
    "项目编码（财务部）",
    "项目状态",
    "实际完成产值",
    "累计收款",
]
MERGED_APPENDIX_NUMERIC_COLUMNS = [
    "实际完成产值",
    "累计收款",
    "应收未收款",
    "预收款应收款",
    "预收款实收款",
    "资金结余",
    "最近一期成本分析利润率",
    "工期延误天数",
]
MERGED_APPENDIX_DATE_COLUMNS = [
    "预收款约定支付日期",
    "预收款实际收款日期",
    "保证金约定退还日期",
    "保证金实际回收日期",
]


def _safe_int(x) -> int:
    """Convert to int, returning 0 for NaN/None/invalid."""
    v = safe_float(x)
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return 0
    return int(v)


def _read_excel_sheet(sheet_pattern: str, data_start_row: int, filepath: str = None) -> pd.DataFrame:
    """Read a sheet from the legacy appendix workbook using the first data row."""
    workbook_path = _resolve_appendix_excel_path(filepath)
    wb = openpyxl.load_workbook(workbook_path, data_only=True)

    for name in wb.sheetnames:
        if sheet_pattern in name:
            ws = wb[name]
            break
    else:
        return pd.DataFrame()

    ncols = ws.max_column
    rows = []
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        values = list(row) if row else []
        if len(values) < ncols:
            values = values + [None] * (ncols - len(values))
        elif len(values) > ncols:
            values = values[:ncols]
        rows.append(values)

    df = pd.DataFrame(rows)
    return df.dropna(how="all").reset_index(drop=True)


def _load_merged_appendix_1(filepath: str) -> pd.DataFrame:
    """Load the new merged appendix/SAP workbook.

    Normalises column-name variants (e.g. 项目编码（DMP）→ 项目编码（财务部）)
    so downstream code sees a consistent interface.
    """
    df = pd.read_excel(filepath, sheet_name=0)
    if df.empty:
        return df

    df = df.dropna(how="all").reset_index(drop=True)

    # ── Normalise project-code column name ──
    _col_rename = {}
    for col in df.columns:
        col_str = str(col).strip()
        if "项目编码" in col_str and "DMP" in col_str:
            _col_rename[col] = "项目编码（财务部）"
    if _col_rename:
        df = df.rename(columns=_col_rename)

    available = {str(col).strip() for col in df.columns}
    # Require at least the project code column; status/value columns are
    # checked individually by the supplement step.
    if "项目编码（财务部）" not in available:
        return pd.DataFrame()

    for col in ["项目编码（财务部）", "项目状态", "未开工或退场原因"]:
        if col in df.columns:
            df[col] = df[col].astype(str).replace("nan", "").str.strip()

    for col in MERGED_APPENDIX_NUMERIC_COLUMNS:
        if col in df.columns:
            df[col] = df[col].apply(safe_float)

    for col in MERGED_APPENDIX_DATE_COLUMNS:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    return df.reset_index(drop=True)


def load_appendix_1(filepath: str = None) -> pd.DataFrame:
    """
    Load Appendix 1.

    Priority:
    1. New merged appendix/SAP workbook.
    2. Legacy project statistics workbook.
    """
    filepath = _resolve_appendix_excel_path(filepath)

    merged_df = _load_merged_appendix_1(filepath)
    if not merged_df.empty:
        return merged_df

    df = _read_excel_sheet("1-项目统计表", data_start_row=7, filepath=filepath)
    if df.empty:
        return df

    mask = df.iloc[:, 2].notna() & (df.iloc[:, 2].astype(str).str.strip() != "")
    df = df[mask].reset_index(drop=True)

    key_positions = {
        "序号（原始）": 0,
        "序号": 1,
        "项目编码（财务部）": 2,
        "项目名称（财务部）": 3,
        "项目名称（市场部）": 4,
        "项目状态": 5,
        "所属公司": 6,
        "营销点名称": 7,
        "是否新兴业务": 8,
        "项目类型": 9,
        "发包方式": 10,
        "承包模式": 11,
        "承接模式": 12,
        "管理模式": 13,
        "项目所在地": 14,
        "所属区域": 15,
        "是否核心区": 16,
        "业主名称": 17,
        "业主性质": 18,
        "是否战略客户": 19,
        "签约法人主体": 20,
        "合同签订时间": 21,
        "合同签订年度": 22,
        "合同类型": 23,
        "合同额(元)": 24,
        "自施合同额(元)": 25,
        "合同约定工期": 26,
        "开工日期": 29,
        "竣工日期": 30,
        "实际开工日期": 31,
        "实际竣工日期": 32,
        "实际工期": 33,
        "工期延误天数": 34,
        "未开工或退场原因": 35,
        "三证取得情况": 36,
        "合作风险分析": 37,
        "进度付款比例": 38,
        "预付款约定": 39,
        "是否三高": 40,
        "是否垫资": 44,
        "付款形式": 45,
        "是否突破营销底线": 46,
        "投标测算利润率": 60,
        "目标责任利润率": 61,
        "最近一期成本分析利润率": 64,
        "实际完成产值": 65,
        "实际发生成本": 66,
        "税金及附加": 67,
        "实际利润": 68,
        "累计收款": 72,
        "累计付款": 73,
        "资金结余": 74,
        "负流原因分析": 75,
        "预收款应收款": 76,
        "预收款实收款": 77,
        "应收未收款": 78,
        "应收未收原因": 79,
    }

    ncols = df.shape[1]
    for name, idx in key_positions.items():
        if idx < ncols:
            df[name] = df.iloc[:, idx]

    if "合同签订年度" in df.columns:
        df["合同签订年度"] = df["合同签订年度"].apply(_safe_int)

    amount_cols = [
        "合同额(元)", "自施合同额(元)", "实际完成产值", "实际发生成本",
        "实际利润", "累计收款", "累计付款", "资金结余",
        "预收款应收款", "预收款实收款", "应收未收款",
        "投标测算利润率", "目标责任利润率", "最近一期成本分析利润率",
        "工期延误天数",
    ]
    for col in amount_cols:
        if col in df.columns:
            df[col] = df[col].apply(safe_float)

    return df.reset_index(drop=True)


def load_appendix_2(filepath: str = None) -> pd.DataFrame:
    """Load Appendix 2: Contract Conversion Rate."""
    df = _read_excel_sheet("2-合同额转化率", data_start_row=5, filepath=filepath)
    if df.empty:
        return df

    col_names = ["序号", "分公司", "年度", "新签合同额", "已签约未转化项目合同额", "确认合同额", "当年完成产值", "合同额转化率", "备注"]
    for i, name in enumerate(col_names):
        if i < df.shape[1]:
            df.rename(columns={df.columns[i]: name}, inplace=True)

    for col in ["新签合同额", "确认合同额", "当年完成产值", "合同额转化率"]:
        if col in df.columns:
            df[col] = df[col].apply(safe_float)
    return df.reset_index(drop=True)


def load_appendix_3(filepath: str = None) -> pd.DataFrame:
    """Load Appendix 3: Bid Win Rate."""
    df = _read_excel_sheet("3-中标率", data_start_row=4, filepath=filepath)
    if df.empty:
        return df
    return df.reset_index(drop=True)


def load_appendix_4(filepath: str = None) -> pd.DataFrame:
    """Load Appendix 4: Marketing Point Info."""
    df = _read_excel_sheet("4-营销点信息表", data_start_row=3, filepath=filepath)
    if df.empty:
        return df
    return df.reset_index(drop=True)


def load_appendix_5(filepath: str = None) -> pd.DataFrame:
    """Load Appendix 5: Contract Target Indicators."""
    df = _read_excel_sheet("5-合同额指标", data_start_row=4, filepath=filepath)
    if df.empty:
        return df
    return df.reset_index(drop=True)


def load_appendix_6(filepath: str = None) -> pd.DataFrame:
    """Load Appendix 6: Strategic Customer Statistics."""
    df = _read_excel_sheet("6-战略客户统计表", data_start_row=5, filepath=filepath)
    if df.empty:
        return df
    return df.reset_index(drop=True)


def load_all_appendices() -> dict:
    """Load all appendix tables. Missing sheets return empty DataFrames."""
    appendices = {}

    print("Loading Appendix 1 (project statistics / merged appendix finance sheet)...")
    appendices["appendix_1"] = load_appendix_1()
    print(f"  -> {len(appendices['appendix_1'])} rows")

    print("Loading Appendix 2 (contract conversion rate)...")
    appendices["appendix_2"] = load_appendix_2()
    print(f"  -> {len(appendices['appendix_2'])} rows")

    print("Loading Appendix 3 (bid win rate)...")
    appendices["appendix_3"] = load_appendix_3()
    print(f"  -> {len(appendices['appendix_3'])} rows")

    print("Loading Appendix 4 (marketing point info)...")
    appendices["appendix_4"] = load_appendix_4()
    print(f"  -> {len(appendices['appendix_4'])} rows")

    print("Loading Appendix 5 (contract targets)...")
    appendices["appendix_5"] = load_appendix_5()
    print(f"  -> {len(appendices['appendix_5'])} rows")

    print("Loading Appendix 6 (strategic customer statistics)...")
    appendices["appendix_6"] = load_appendix_6()
    print(f"  -> {len(appendices['appendix_6'])} rows")

    return appendices
